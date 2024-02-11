import math

import numpy as np
import os
import typing

from typing import TYPE_CHECKING

from source.base.metrics import get_metric_meshes

if TYPE_CHECKING:
    import pandas as pd


def make_excel_file_comparison(cd_pred_list, human_readable_results, output_file, val_set,
                               low_metrics_better: typing.Union[bool, typing.Sequence] = True):
    import pandas as pd
    # try https://realpython.com/openpyxl-excel-spreadsheets-python/

    # one shape per line, dataset per column
    cd_pred = np.array(cd_pred_list).transpose()
    data_headers_human_readable = ['Shape'] + [hr for hr in human_readable_results]
    # data_headers = [''] + [rft for rft in result_file_templates]
    data_body = [[val_set[i]] + cd_pred[i].tolist() for i in range(len(val_set))]
    df = pd.DataFrame(data=data_body, columns=data_headers_human_readable)
    df = df.set_index('Shape')

    export_xlsx(df=df, low_metrics_better=low_metrics_better, output_file=output_file, add_stats=True,
                header=True, independent_cols=True)


def make_quantitative_comparison(
        shape_names: typing.Sequence[str], gt_mesh_files: typing.Sequence[str],
        result_headers: typing.Sequence[str], result_file_templates: typing.Sequence[str],
        comp_output_dir: str, num_samples=10000, num_processes=0):

    cd_pred_list = get_metric_meshes(
        result_file_template=result_file_templates, shape_list=shape_names, gt_mesh_files=gt_mesh_files,
        num_samples=num_samples, metric='chamfer', num_processes=num_processes)
    cd_output_file = os.path.join(comp_output_dir, 'chamfer_distance.xlsx')
    make_excel_file_comparison(cd_pred_list, result_headers, cd_output_file, shape_names, low_metrics_better=[True])

    f1_pred_list = get_metric_meshes(
        result_file_template=result_file_templates, shape_list=shape_names, gt_mesh_files=gt_mesh_files,
        num_samples=num_samples, metric='f1', num_processes=num_processes)
    cd_output_file = os.path.join(comp_output_dir, 'f1.xlsx')
    make_excel_file_comparison(f1_pred_list, result_headers, cd_output_file, shape_names, low_metrics_better=[False])

    iou_pred_list = get_metric_meshes(
        result_file_template=result_file_templates, shape_list=shape_names,
        gt_mesh_files=gt_mesh_files, num_samples=num_samples, metric='iou', num_processes=num_processes)
    iou_output_file = os.path.join(comp_output_dir, 'iou.xlsx')
    make_excel_file_comparison(iou_pred_list, result_headers, iou_output_file, shape_names, low_metrics_better=[False])

    nc_pred_list = get_metric_meshes(
        result_file_template=result_file_templates, shape_list=shape_names,
        gt_mesh_files=gt_mesh_files, num_samples=num_samples, metric='normals', num_processes=num_processes)
    nc_output_file = os.path.join(comp_output_dir, 'normal_error.xlsx')
    make_excel_file_comparison(nc_pred_list, result_headers, nc_output_file, shape_names, low_metrics_better=[True])


def make_html_report(report_file_out, comp_name, pc_renders, gt_renders,
                     cd_vis_renders, dist_cut_off, metrics_cd, metrics_iou, metrics_nc):

    num_rows = len(gt_renders)
    num_recs = len(metrics_cd)
    num_cols = len(metrics_cd) + 3

    def clean_path_list(path_list: typing.Sequence[str]):
        return [p.replace('\\', '/') for p in path_list]

    gt_renders = clean_path_list(gt_renders)
    pc_renders = clean_path_list(pc_renders)
    cd_vis_renders = [clean_path_list(rec_shapes) for rec_shapes in cd_vis_renders]

    def make_relative(path_list: typing.Sequence[str]):
        import pathlib
        return [pathlib.Path(*pathlib.Path(p).parts[3:]) for p in path_list]

    gt_renders = make_relative(gt_renders)
    pc_renders = make_relative(pc_renders)
    cd_vis_renders = [make_relative(rec_shapes) for rec_shapes in cd_vis_renders]

    def make_human_readable(path_list):
        return [str(p).replace('_', ' ') for p in path_list]

    shape_names_hr = make_human_readable([os.path.basename(p) for p in gt_renders])
    rec_names = make_human_readable([os.path.split(os.path.dirname(os.path.dirname(cd_vis_renders[ir][0])))[1]
                                     for ir in range(num_recs)])
    gt_renders_hr = make_human_readable(gt_renders)
    pc_renders_hr = make_human_readable(pc_renders)
    cd_vis_renders_hr = [make_human_readable(rec_shapes) for rec_shapes in cd_vis_renders]

    # template draft by Chat-GPT
    html_template = r'''
<!DOCTYPE html>
<html>
  <head>
    <title>PPSurf Comparison Results</title>
    <style>
      .hide-column {{
        display: none;
      }}
      tr:hover {{
        background-color: #D6EEEE;
      }}
      td {{
        border: 1px solid #000;
      }}
      table {{width: 100%}}
      th {{
        border: 1px solid #000000;
        background-color: #EEEEEE;
        width: {th_width}%;
        position: sticky;
        top: 0;
        z-index:2;
      }}
      tr {{vertical-align: top}}
      th.sticky, td.sticky {{
        position: sticky;
        left: 0;
      }}
      td.sticky {{
        z-index:1;
      }}
    </style>
    <script>
      var numColumns = 2 + {num_rec}; // number of columns in the table

      function toggleColumn(columnIndex) {{
        var table = document.getElementById("image-table");
        var columnCells = table.getElementsByTagName("td");
        for (var i = 0; i < columnCells.length; i += numColumns) {{
          if (i + columnIndex < columnCells.length) {{
            var cell = columnCells[i + columnIndex];
            if (cell.classList.contains("hide-column")) {{
              cell.classList.remove("hide-column");
            }} else {{
              cell.classList.add("hide-column");
            }}
          }}
        }}
      }}
    </script>
  </head>
  <body>
    <h1>Dataset: {title}</h1>
    <table>
      <thead>
        <tr>
{th_rec}
        </tr>
      </thead>
      <tbody>
{tr}
      </tbody>
    </table>
  </body>
</html>
    '''

    table_header_rec_template = """
          <th{th_sticky}>{rec_name}</th>"""
    table_row_template = '''
        <tr>
          <td class="sticky">{file_name}</td>
          <td class="sticky"><img src="{pc_file}" alt="{pc_file_hr}" width="{img_size}" height="{img_size}"></td>
          <td class="sticky"><img src="{gt_file}" alt="{gt_file_hr}" width="{img_size}" height="{img_size}"></td>
{recs}
        </tr>'''
    table_row_rec_template = '          <td><img src="{rec_file}" alt="{rec_file_hr}"'\
                             'width="{img_size}" height="{img_size}"><br>{metrics}</td>'
    table_row_rec_metrics_template = 'CD: {cd:.2f}, IoU: {iou:.2f}, NCE: {nc:.2f}'

    img_size = 300
    table_row_rec_metrics = [[table_row_rec_metrics_template.format(
        cd=metrics_cd[ir][i] * 100.0, iou=metrics_iou[ir][i], nc=metrics_nc[ir][i])
        for i in range(num_rows)]
        for ir in range(num_recs)]
    table_row_rec = [[table_row_rec_template.format(
        rec_file=cd_vis_renders[ir][i], rec_file_hr=cd_vis_renders_hr[ir][i],
        metrics=table_row_rec_metrics[ir][i], img_size=img_size)
        for ir in range(num_recs)]
        for i in range(num_rows)]

    table_rows = [table_row_template.format(
        file_name=shape_names_hr[i],
        pc_file=pc_renders[i], pc_file_hr=pc_renders_hr[i],
        gt_file=gt_renders[i], gt_file_hr=gt_renders_hr[i],
        recs='\n'.join(table_row_rec[i]), img_size=img_size)
        for i in range(num_rows)]

    th_width = int(math.floor(100 / num_cols))
    th_names = ['Shape Name', 'Point Cloud', 'GT Object'] + rec_names
    th_sticky = [' class="sticky"'] * 3 + [''] * len(rec_names)
    table_header_rec = ''.join([table_header_rec_template.format(th_sticky=th_sticky[ni], rec_name=n)
                                for ni, n in enumerate(th_names)])

    html_text = html_template.format(
        th_width=th_width,
        num_rec=num_recs, title=comp_name,
        th_rec=table_header_rec, tr=''.join(table_rows))

    with open(report_file_out, 'w') as text_file:
        text_file.write(html_text)


def make_test_report(shape_names: list, results: typing.Union[list, dict],
                     output_file: str, output_names: list, is_dict=True):
    import pandas as pd
    from torch import stack

    metrics_keys_to_log = ['abs_dist_rms', 'accuracy', 'precision', 'recall', 'f1_score']
    headers = ['Shape', 'Loss total'] + output_names + metrics_keys_to_log
    low_metrics_better = [True] * (1 + len(output_names)) + [True, False, False, False, False]

    if not is_dict:
        loss_total = [r[0] for r in results]
        loss_components = [r[1] for r in results]
        metrics_dicts = [r[2] for r in results]
        metrics_lists = []
        for m in metrics_keys_to_log:
            metrics_list = [md[m] for md in metrics_dicts]
            metrics_lists.append(metrics_list)
        metrics = np.array(metrics_lists).transpose()
    else:
        loss_total = results['loss'].detach().cpu()
        loss_components = results['loss_components_mean'].detach().cpu()
        if len(loss_components.shape) == 1:
            loss_components = loss_components.unsqueeze(1)
        metrics = stack([results[k] for k in metrics_keys_to_log]).transpose(0, 1).detach().cpu()

        if len(loss_total.shape) == 2:  # DP -> squeeze
            loss_total = loss_total.squeeze(-1)
            metrics = metrics.squeeze(-1)

    data = [[shape_names[i]] + [loss_total[i].item()] + loss_components[i].tolist() + metrics[i].tolist()
            for i in range(len(loss_total))]
    df = pd.DataFrame(data=data, columns=headers)
    df = df.set_index('Shape')

    export_xlsx(df=df, low_metrics_better=low_metrics_better, output_file=output_file, 
                add_stats=True, header=True, independent_cols=True)

    loss_total_mean = np.mean(np.array(loss_total))
    abs_dist_rms_mean = np.nanmean(metrics[:, 0])
    f1_mean = np.nanmean(metrics[:, -1])
    return loss_total_mean, abs_dist_rms_mean, f1_mean


def export_xlsx(df: 'pd.DataFrame', low_metrics_better: typing.Union[None, typing.Sequence[bool], bool],
                output_file: str, add_stats=True, header=True, independent_cols=True):
    import datetime
    from source.base import fs

    # export with conditional formatting and average
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    wb = Workbook()
    ws = wb.active

    df_export = df.copy()
    df_export.reset_index(inplace=True)  # revert index to normal column to get rid of extra header row
    for r in dataframe_to_rows(df_export, index=False, header=header):
        ws.append(r)

    # no direction given, assume near 0 or near 1 results
    if low_metrics_better is None:
        cols = df.to_numpy()
        cols = np.vectorize(lambda x: x.timestamp() if isinstance(x, datetime.datetime) else x)(cols)
        cols_mean = np.nanmean(cols, axis=0)
        if not independent_cols:
            cols_mean = np.mean(cols_mean)  # scalar for dependent cols
        low_metrics_better = np.logical_or(cols_mean > 1.0, cols_mean < 0.5)

    top_row = 2
    col_ids = df.index.shape[1] if len(df.index.shape) > 1 else 1
    ws.freeze_panes = '{}{}'.format(get_column_letter(col_ids + 1), top_row)
    bottom_row = df.shape[0] + top_row - 1
    if add_stats:
        for di in range(df.shape[1]):
            column = col_ids + 1 + di
            column_letter = get_column_letter(column)
            ws.cell(row=bottom_row + 1, column=column).value = '=AVERAGE({}{}:{}{})'.format(
                column_letter, top_row, column_letter, bottom_row)
            ws.cell(row=bottom_row + 2, column=column).value = '=MEDIAN({}{}:{}{})'.format(
                column_letter, top_row, column_letter, bottom_row)
            # ws.cell(row=bottom_row + 3, column=column).value = '=STDEV.P({}{}:{}{})'.format(  # strange '@' appears
            ws.cell(row=bottom_row + 3, column=column).value = '=STDEV({}{}:{}{})'.format(  # rely on compatibility
                column_letter, top_row, column_letter, bottom_row)

        # Stat names
        ws.cell(row=bottom_row + 1, column=1).value = 'AVERAGE'
        ws.cell(row=bottom_row + 2, column=1).value = 'MEDIAN'
        ws.cell(row=bottom_row + 3, column=1).value = 'STDEV'

    def add_formatting_rule(col_start_id, row_start_id, col_end_id, row_end_id, lower_is_better):
        col_start_letter = get_column_letter(col_start_id)
        col_end_letter = get_column_letter(col_end_id)
        col_range_str = '{col_start}{row_start}:{col_end}{row_end}'.format(
            col_start=col_start_letter, row_start=row_start_id, col_end=col_end_letter, row_end=row_end_id)
        if lower_is_better:  # error here means that this list has an invalid length
            start_color = 'FF00AA00'
            end_color = 'FFAA0000'
        else:
            end_color = 'FF00AA00'
            start_color = 'FFAA0000'
        rule = ColorScaleRule(start_type='percentile', start_value=0, start_color=start_color,
                              mid_type='percentile', mid_value=50, mid_color='FFFFFFFF',
                              end_type='percentile', end_value=100, end_color=end_color)
        ws.conditional_formatting.add(col_range_str, rule)

        # highlight optimum
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Font

        asc_desc = 'MIN' if lower_is_better else 'MAX'
        # should be like =H2=MIN(H$2:H$11)
        formula = '={col_start}{row_start}={func}({col_start}${row_start}:{col_end}${row_end})'.format(
            col_start=col_start_letter, row_start=row_start_id, func=asc_desc,
            col_end=col_end_letter, row_end=row_end_id)
        rule = FormulaRule(formula=(formula,), font=Font(underline='single'))
        ws.conditional_formatting.add(col_range_str, rule)

    # color scale over shapes
    if independent_cols:
        bottom_row_formatting = bottom_row + (2 if add_stats else 0)  # not for STDEV
        for col in range(df.shape[1]):
            if not np.isnan(low_metrics_better[col]):
                add_formatting_rule(col_start_id=col+col_ids+1, row_start_id=top_row,
                                    col_end_id=col+col_ids+1, row_end_id=bottom_row_formatting,
                                    lower_is_better=low_metrics_better[col])
    else:  # dependent cols
        for shape_id in range(df.shape[0]):
            row = top_row + shape_id
            add_formatting_rule(col_start_id=col_ids+1, row_start_id=row,
                                col_end_id=df.shape[1]+col_ids+1, row_end_id=row,
                                lower_is_better=low_metrics_better)

        # color scale over stats (horizontal)
        lower_better = [low_metrics_better] * 2 + [True]  # lower stdev is always better, mean and avg depend on metric
        for stat_id in range(3):
            row = bottom_row + 1 + stat_id
            add_formatting_rule(col_start_id=col_ids+1, row_start_id=row,
                                col_end_id=df.shape[1] + col_ids+1, row_end_id=row,
                                lower_is_better=lower_better[stat_id])

    fs.make_dir_for_file(output_file)
    wb.save(output_file)


def _drop_stats_rows(df: 'pd.DataFrame',
                     stats: typing.Sequence[str] = ('AVG', 'AVERAGE', 'MEAN', 'MEDIAN', 'STDEV.P', 'STDEV'))\
        -> 'pd.DataFrame':
    df = df.copy()
    for stat in stats:
        df = df.drop(stat, errors='ignore')
    return df


def make_dataset_comparison(results_reports: typing.Sequence[typing.Sequence[str]], output_file: str):
    import time
    import pandas as pd

    def _get_header_and_mean(report_file: typing.Union[str, typing.Any]):
        metrics_type = os.path.basename(report_file)
        metrics_type = os.path.splitext(metrics_type)[0]

        if not os.path.isfile(report_file):
            method_name = os.path.basename(os.path.split(os.path.split(report_file)[0])[0])
            headers = ['Model', 'Mean {}'.format(metrics_type),
                       'Median {}'.format(metrics_type), 'Stdev {}'.format(metrics_type), ]
            data = [method_name, np.nan, np.nan, np.nan, ]

            df_missing = pd.DataFrame(data=[data], columns=headers)
            df_missing = df_missing.set_index('Model')
            return df_missing

        df = pd.read_excel(io=report_file, header=0, index_col=0)
        df = _drop_stats_rows(df)

        if len(df.columns) == 1:  # CD, IoU, NC -> single columns in multiple files
            df_name = df.columns[0]
            df_mean = df.mean(axis=0)[0]
            df_median = df.median(axis=0)[0]
            df_stdev = df.std(axis=0)[0]
            headers = ['Model', 'Mean {}'.format(metrics_type),
                       'Median {}'.format(metrics_type), 'Stdev {}'.format(metrics_type), ]
            data = [df_name, df_mean, df_median, df_stdev, ]

            df_means = pd.DataFrame(data=[data], columns=headers)
            df_means = df_means.set_index('Model')
        else:  # Test, only one file with multiple columns
            series_means = df.mean(axis=0)
            df_means: pd.DataFrame = series_means.to_frame().transpose()
            model_name = os.path.basename(report_file).split('metrics_')[1]
            model_name = os.path.splitext(model_name)[0]
            df_means.insert(0, 'Model', [model_name])
            df_means = df_means.set_index('Model')
            test_time = np.datetime64(time.strftime('%Y-%m-%dT%H:%M:%S', time.gmtime(os.path.getmtime(report_file))))
            df_means.insert(0, 'Date', [test_time])
            df_means.insert(0, 'Count', [float(df.shape[0])])  # float to keep the array dtype
        return df_means

    def assemble_model_data(reports_model: typing.Sequence[str]):
        df_model_list = [_get_header_and_mean(f) for f in reports_model]
        df_model = pd.concat(df_model_list, axis=1)
        return df_model

    df_mean_list = [assemble_model_data(l) for l in results_reports]
    df_mean_all = pd.concat(df_mean_list, axis=0)

    # df_mean_all.insert(0, 'Model', [str(f) for f in results_reports])
    # df_mean_all = df_mean_all.set_index('Model')

    # df_mean_all = df_mean_all.sort_values('f1_score', ascending=False)
    df_mean_all = df_mean_all.sort_values('Mean chamfer_distance', ascending=False)
    export_xlsx(df=df_mean_all, low_metrics_better=None, output_file=output_file, add_stats=False,
                header=True, independent_cols=True)


def assemble_quantitative_comparison(
        comp_output_dir: str,
        report_path_templates=('results/poco_blensor_prec32_again/{}.xlsx',),
        metrics=('chamfer_distance', 'iou', 'normal_error', 'f1'),
        metrics_lower_better=(True, False, True, False)):
    import pandas as pd

    def assemble_report(report_paths: typing.Sequence[str]):
        reports_list = []
        for p in report_paths:
            if not os.path.isfile(p):
                print('Missing report: {}'.format(p))

                model_name = os.path.split(os.path.split(os.path.dirname(p))[0])[1]
                headers = ['Shape', model_name]
                df_report = pd.DataFrame(columns=headers)
                df_report = df_report.set_index('Shape')
                reports_list.append(df_report)
            else:
                df_report: pd.DataFrame = pd.read_excel(io=p, header=0, index_col=0)
                reports_list.append(df_report)

        df = pd.concat(reports_list, axis=1)
        df = _drop_stats_rows(df)
        return df

    results_per_shape_dict = {}
    for mi, m in enumerate(metrics):
        report_paths_metric = [t.format(m) for t in report_path_templates]
        df_m = assemble_report(report_paths_metric)
        results_per_shape_dict[m] = df_m.to_numpy()

        report_file = os.path.join(comp_output_dir, '{}.xlsx'.format(m))
        export_xlsx(df=df_m, low_metrics_better=metrics_lower_better[mi], output_file=report_file, add_stats=True,
                    header=True, independent_cols=False)

    return results_per_shape_dict


def _prettify_df(df: 'pd.DataFrame'):
    import re

    def _replace_case_insensitive(text, word, replacement):
        regex_ignore_case = '(?i){}'.format(word)
        return re.sub(regex_ignore_case, replacement, text)

    # keep index title in row with CD, IoU, NC
    df.reset_index(inplace=True)

    find_replace = (
        ('abc', 'ABC'),
        ('famous', 'Famous'),
        ('thingi10k_scans', 'Thingi10k'),
        ('chamfer_distance', 'Chamfer Distance (x100)'),
        ('iou', 'IoU'),
        ('normal_error', 'Normal Error'),
        ('normal_consistency', 'Normal Error'),
        ('sap', 'SAP'),
        ('p2s', 'P2S'),
        ('poco', 'POCO'),
        ('pts_gen_sub3k_iter10', ''),
        ('ppsurf', 'PPSurf'),
        ('_vanilla_zeros_global', ' only local'),
        ('_vanilla_zeros_local', ' only global'),
        ('_vanilla_sym_max', ' sym max'),
        ('_vanilla_qpoints', ' qpoints'),
        ('_vanilla', ' merge cat'),
        ('_merge_sum', ' merge sum'),
        ('optim', 'O'),
        ('mean ', ''),
        ('_', ' '),
    )
    for fr in find_replace:
        # replace in values
        regex_ignore_case_values = '(?i){}'.format(fr[0])
        df.replace(to_replace=regex_ignore_case_values, value=fr[1], inplace=True, regex=True)

        # rename in multi
        for multi_col in df.columns:
            col: str  # type annotation not allowed in for
            for col in multi_col:
                if col.lower().find(fr[0].lower()) >= 0:
                    df.rename(columns={col: _replace_case_insensitive(col, fr[0], fr[1])}, inplace=True)

    # factor 100 for chamfer distance
    cols_chamfer = [c for c in df.columns if c[0].find('Chamfer Distance') >= 0]
    for c in cols_chamfer:
        df[c] = df[c] * 100

    return df


def xslx_to_latex(xlsx_file: str, latex_file: str):
    import pandas as pd

    df: pd.DataFrame = pd.read_excel(io=xlsx_file, header=0, index_col=0)

    # nicer column names
    columns = [
        ('Chamfer Distance (x100)', 'Mean'),
        ('Chamfer Distance (x100)', 'Median'),
        ('Chamfer Distance (x100)', 'Stdev'),
        ('IoU', 'Mean'),
        ('IoU', 'Median'),
        ('IoU', 'Stdev'),
        ('F1', 'Mean'),
        ('F1', 'Median'),
        ('F1', 'Stdev'),
        ('Normal Error', 'Mean'),
        ('Normal Error', 'Median'),
        ('Normal Error', 'Stdev'),
    ]
    df.columns = pd.MultiIndex.from_tuples(columns)

    df = _prettify_df(df)
    df.to_latex(buf=latex_file, float_format='%.2f', na_rep='-', index=False, bold_rows=True,
                column_format='l' + 'c' * (df.shape[1] - 1), escape=False)

    # strange stuff with styler. why can't I give this to the df for export?
    # styler = df.style.highlight_max(axis=None, props='font-weight:bold;', subset=columns)
    # styler.format('{:.2f}', na_rep='-', subset=columns)
    # styler.to_latex(buf=latex_file, column_format='l' + 'c' * (df.shape[1] - 1))


def merge_comps(comp_list: typing.Sequence[str], comp_merged_out_file: str,
                comp_merged_out_latex: str, methods_order: typing.Optional[list], float_format='%.2f'):
    import pandas as pd
    comp_list_existing = [f for f in comp_list if os.path.isfile(f)]
    if len(comp_list_existing) == 0:
        print('WARNING: No metrics found for: {}'.format(comp_list))
        return
    
    dfs = [pd.read_excel(io=f, header=0, index_col=0) for f in comp_list_existing]
    datasets = [os.path.split(os.path.dirname(f))[1] for f in comp_list_existing]
    dfs_with_ds = [df.assign(dataset=datasets[dfi]) for dfi, df in enumerate(dfs)]
    dfs_multiindex = [df.set_index(['dataset', df.index]).T for df in dfs_with_ds]

    def _extract_metric(df_list: typing.Sequence[pd.DataFrame], order: typing.Optional[list], metric: str):
        df_metric = [df.xs(metric, axis=0) for df in df_list]

        # dataset name as index, metric as column
        df_metric_id = [df.reset_index(level=0) for df in df_metric]
        df_metric_for_col = [df.rename(columns={metric: datasets[dfi]}) for dfi, df in enumerate(df_metric_id)]
        df_metric_for_col = [df.drop(columns=['dataset']).T for df in df_metric_for_col]

        # xs removes the column name, so we need to add it again
        # df_metric_ds_col = [df.rename(index=datasets[dfi]).T for dfi, df in enumerate(df_metric)]

        df_metric_merged = pd.concat(df_metric_for_col, axis=0)

        if order is not None and len(order) > 0:
            df_metric_merged = df_metric_merged[order]

        df_metric_merged_with_ds = df_metric_merged.T.assign(metric=metric)
        df_metric_merged_id = df_metric_merged_with_ds.set_index(['metric', df_metric_merged_with_ds.index]).T

        return df_metric_merged_id

    df_cd =  _extract_metric(df_list=dfs_multiindex, order=methods_order, metric='Mean chamfer_distance')
    df_iou = _extract_metric(df_list=dfs_multiindex, order=methods_order, metric='Mean iou')
    df_f1 =  _extract_metric(df_list=dfs_multiindex, order=methods_order, metric='Mean f1')
    df_nc =  _extract_metric(df_list=dfs_multiindex, order=methods_order, metric='Mean normal_error')

    df_merged: pd.DataFrame = pd.concat((df_cd, df_iou, df_f1, df_nc), axis=1)

    # add mean row
    df_mean_row = df_merged.mean(axis=0).rename('Mean')
    df_merged = pd.concat((df_merged, pd.DataFrame(df_mean_row).T), axis=0)

    df_merged = _prettify_df(df_merged)
    df_merged.rename(columns={'index': 'Dataset'}, inplace=True)

    from source.base.fs import make_dir_for_file
    make_dir_for_file(comp_merged_out_file)
    df_merged.to_excel(comp_merged_out_file, float_format=float_format)
    make_dir_for_file(comp_merged_out_latex)
    # TODO: to_latex is deprecated, use df.style.to_latex instead
    df_merged.to_latex(buf=comp_merged_out_latex, float_format=float_format, na_rep='-', index=False, bold_rows=True,
                       column_format='l' + 'c' * (df_merged.shape[1] - 1), escape=False)

